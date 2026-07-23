"""
DCN Calculation Module - Excel export (client pricing sheet).

v1 produces a three-sheet workbook, in memory (never written server-side):

  Pricing  - client-facing: the package hierarchy of the master with SELL
             prices (cost + levies scaled by the revision's markup waterfall,
             applied uniformly so packages sum exactly to the grand total).
  Elements - internal: element subtotals (the seven elementen) per package
             plus the master rollup - the IBIS-style check view.
  Basis    - provenance: Q number, revision, rate-set label, fx snapshot,
             markup percentages, export timestamp.

The client-facing layout will be refined against a real client template with
Patrick; the numbers and structure are final, the styling is v1.
"""
import io
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers

from app.calcmod import repo, engine
from app.calcmod.db import ELEMENTS, ELEMENT_LABELS, SPLIT_ELEMENTS

TEAL = "0F766E"
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor=TEAL)
BOLD = Font(bold=True)
MONEY = numbers.FORMAT_CURRENCY_USD_SIMPLE


def _sell_factor(res, master_id):
    m = res["blocks"][master_id]
    base = m["cost"] + m["levies"]
    return (m["sell"] / base) if base else 1.0


def workbook_bytes(qnumber, rev_no=None):
    calc = repo.get_calc(qnumber)
    rev = repo.get_revision(qnumber, rev_no)
    tree = repo.get_tree(rev["id"])
    snap = repo.load_snapshot(rev["id"])
    res = engine.compute(tree, snap)
    master_id = res["master_id"]
    factor = _sell_factor(res, master_id)

    blocks = {b["id"]: b for b in tree["blocks"]}
    kids = {}
    for b in tree["blocks"]:
        if b["parent_id"] and b["kind"] == "package":
            kids.setdefault(b["parent_id"], []).append(b)

    wb = Workbook()

    # ---- Pricing ----------------------------------------------------------
    ws = wb.active
    ws.title = "Pricing"
    ws["A1"] = f"{qnumber} rev {rev['rev_no']} \u2014 {calc['title']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = calc.get("client") or ""
    ws.append([]); ws.append(["Item", "Unit", "Price (USD)"])
    for cell in ws[4]:
        cell.font, cell.fill = HDR, HDR_FILL
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    def walk(bid, depth):
        for b in sorted(kids.get(bid, []), key=lambda x: (x["sort_order"], x["id"])):
            r = res["blocks"][b["id"]]
            sell = (r["cost"] + r["levies"]) * factor
            ws.append(["    " * depth + b["name"], b["unit_label"], round(sell, 2)])
            ws.cell(ws.max_row, 3).number_format = MONEY
            if depth == 0:
                ws.cell(ws.max_row, 1).font = BOLD
                ws.cell(ws.max_row, 3).font = BOLD
            walk(b["id"], depth + 1)

    walk(master_id, 0)
    m = res["blocks"][master_id]
    own = m["cost"] + m["levies"] - sum(
        (res["blocks"][b["id"]]["cost"] + res["blocks"][b["id"]]["levies"])
        for b in kids.get(master_id, []))
    if own > 0.005:                       # master-level lines / referenced blocks
        ws.append(["General / day-rate items", "", round(own * factor, 2)])
        ws.cell(ws.max_row, 3).number_format = MONEY
    ws.append([])
    ws.append(["TOTAL", "", round(m["sell"], 2)])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.cell(ws.max_row, 3).font = Font(bold=True, size=12)
    ws.cell(ws.max_row, 3).number_format = MONEY

    # ---- Elements ---------------------------------------------------------
    we = wb.create_sheet("Elements")
    split_cols = [f"{ELEMENT_LABELS[e]} {o}" for e in SPLIT_ELEMENTS
                  for o in ("int.", "ext.")]
    we.append(["Package"] + [ELEMENT_LABELS[e] for e in ELEMENTS] + split_cols
              + ["Levies", "Cost total"])
    for cell in we[1]:
        cell.font, cell.fill = HDR, HDR_FILL
    we.column_dimensions["A"].width = 40

    def walk_el(bid, depth):
        for b in sorted(kids.get(bid, []), key=lambda x: (x["sort_order"], x["id"])):
            r = res["blocks"][b["id"]]
            we.append(["    " * depth + b["name"]]
                      + [round(r["elements"][e], 2) for e in ELEMENTS]
                      + [round(r["splits"][e][o], 2) for e in SPLIT_ELEMENTS
                         for o in ("internal", "external")]
                      + [round(r["levies"], 2), round(r["cost"] + r["levies"], 2)])
            walk_el(b["id"], depth + 1)

    walk_el(master_id, 0)
    we.append(["MASTER TOTAL"] + [round(m["elements"][e], 2) for e in ELEMENTS]
              + [round(m["splits"][e][o], 2) for e in SPLIT_ELEMENTS
                 for o in ("internal", "external")]
              + [round(m["levies"], 2), round(m["cost"] + m["levies"], 2)])
    for cell in we[we.max_row]:
        cell.font = BOLD
    for row in we.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = MONEY

    # ---- Basis ------------------------------------------------------------
    wbs = wb.create_sheet("Basis")
    mk = snap["markups"] or {}
    rows = [("Q number", qnumber), ("Revision", rev["rev_no"]),
            ("Status", rev["status"]), ("Division", calc["division"]),
            ("Region", calc["region"]),
            ("Rate set (snapshot)", rev.get("rate_set_label") or "-"),
            ("Exported", datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"),
            ("", ""),
            ("Levy local", mk.get("levy_local_pct", 0)),
            ("Levy expat", mk.get("levy_expat_pct", 0)),
            ("Overhead", mk.get("overhead_pct", 0)), ("Risk", mk.get("risk_pct", 0)),
            ("Profit", mk.get("profit_pct", 0)), ("Margin", mk.get("margin_pct", 0)),
            ("", "")] + [(f"FX {cur} \u2192 USD", r) for cur, r in sorted(snap["fx"].items())]
    for a, b in rows:
        wbs.append([a, b])
    wbs.column_dimensions["A"].width = 26

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def excel_filename(qnumber, rev_no):
    return f"{qnumber}_rev{rev_no}_pricing.xlsx"


# --------------------------------------------------------------------------- #
# Library overview export
# --------------------------------------------------------------------------- #
def library_workbook_bytes(rate_set_id, rate_set_label="", filters=None):
    """The unified library overview as a workbook: one row per item, current
    filters applied (also noted in the header), rates from the given set."""
    from app.calcmod.repo import list_items, code_label

    LIB_LABEL = {"personnel": "Personnel", "equipment": "Equipment",
                 "materials": "Materials", "subcontracting": "Sub-contracting"}
    rows = []
    for lib in ("personnel", "equipment", "materials", "subcontracting"):
        for i in list_items(lib, with_rates_for=rate_set_id):
            rows.append({
                "code": i["code"], "label": code_label(i["code"]),
                "description": i.get("description") or i.get("function") or "",
                "category": LIB_LABEL[lib], "subcat": i.get("category") or "",
                "ownership": i.get("ownership") or "internal",
                "region": i.get("region") or "ALL", "unit": i.get("unit") or "",
                "erp_no": i.get("erp_no") or "",
                "currency": i.get("currency") or "",
                "office_rate": i.get("office_rate"), "yard_rate": i.get("yard_rate"),
                "offshore_rate": i.get("offshore_rate"), "rate": i.get("rate"),
            })
    for f, v in (filters or {}).items():
        rows = [r for r in rows if str(r.get(f) or "") == v]
    rows.sort(key=lambda r: r["code"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Library"
    ws["A1"] = "DCN Calculation Module \u2014 library overview"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Rate set: {rate_set_label or rate_set_id} \u00b7 exported "
                + datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"
                + ("  \u00b7 filters: " + ", ".join(f"{k}={v}" for k, v in
                                                     (filters or {}).items())
                   if filters else ""))
    ws["A2"].font = Font(color="6B7280")
    heads = ["Code", "Code (long)", "Description", "Category", "Sub-category",
             "Int/Ext", "Region", "Unit", "ERP no", "Currency",
             "Office", "Yard", "Offshore", "Rate"]
    ws.append([]); ws.append(heads)
    for cell in ws[4]:
        cell.font, cell.fill = HDR, HDR_FILL
    widths = (14, 24, 34, 16, 18, 9, 8, 8, 14, 9, 10, 10, 10, 12)
    for col, w in zip("ABCDEFGHIJKLMN", widths):
        ws.column_dimensions[col].width = w
    for r in rows:
        ws.append([r["code"], r["label"], r["description"], r["category"],
                   r["subcat"], r["ownership"], r["region"], r["unit"],
                   r["erp_no"], r["currency"], r["office_rate"], r["yard_rate"],
                   r["offshore_rate"], r["rate"]])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

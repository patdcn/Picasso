"""
Calculation - IBIS engine.

Pure-Python port of the former client-side IBIS review tool. An IBIS ``.xtb``
file is a SQLite database; this module parses it into a plain-Python model,
rolls costs up the chapter/line tree, applies the compounded *staart*
(overhead / profit / CAR), computes a location/asset-scoped levy + VAT quote,
exports that quote to Excel, and writes amendments back into a NEW ``.xtb`` copy
(the original bytes are never mutated).

Nothing here touches Dash, the filesystem or the network: everything works on
``bytes`` in / ``bytes`` out so the page layer can keep the uploaded file in a
per-session store and never persist it. That keeps a confidential cost/margin
file transient - it is parsed in memory for the request and discarded.

Terminology note (IBIS is Dutch): a Kostenpost's ``Netto*`` field is the *cost*
and ``Bruto*`` is the *sell* (cost x (1 + line markup)). The four cost types are
Arbeid=Personnel, Materieel=Equipment, Materiaal=Material,
Onderaanneming=Subcontract.

Validated against real files: ``clean`` v153 -> cost 1693.52 / markup 169.35 /
staart 0 / net 1862.87, and ``test`` v8 -> cost 30,315,214.65 / markup
130,074.60 / staart 7.4949% (2,281,850.07) / net 32,727,139.33, each matching
the file's own stored Begroting totals to the cent.
"""
from __future__ import annotations

import io
import os
import sqlite3
import tempfile
from datetime import datetime, timezone

# --------------------------------------------------------------------------- #
# constants
# --------------------------------------------------------------------------- #
# (netto-col, bruto-col, display type, category key)
CATS = [
    ("NettoArbeid", "BrutoArbeid", "Personnel", "Arbeid"),
    ("NettoMaterieel", "BrutoMaterieel", "Equipment", "Materieel"),
    ("NettoMateriaal", "BrutoMateriaal", "Material", "Materiaal"),
    ("NettoOnderaanneming", "BrutoOnderaanneming", "Subcontract", "Onderaanneming"),
]
CATKEYS = ["Arbeid", "Materieel", "Materiaal", "Onderaanneming"]
QLBL = {"Arbeid": "Personnel", "Materieel": "Equipment",
        "Materiaal": "Material", "Onderaanneming": "Subcontract"}
# staart Kostensoort code -> category, used when seeding levies from a file staart
QKMAP = {"ARB": "Arbeid", "MTA": "Materiaal", "MTE": "Materieel", "ODA": "Onderaanneming"}


def r2(x):
    return None if x is None else round(x * 100) / 100


def r5(x):
    return None if x is None else round(x * 1e5) / 1e5


# --------------------------------------------------------------------------- #
# bid presets (formerly baked into the HTML)
# --------------------------------------------------------------------------- #
def bid_presets():
    """Named levy stacks used as one-click bid presets. Kept here rather than in
    the DB so the engine is self-contained; the page can override from params."""
    return {
        "ph": {"name": "Philippines", "levies": [
            {"name": "VAT", "origin": "any", "asset": "any", "rate": 7.5, "per": {}, "vat": True}]},
        "ng_local": {"name": "Nigeria - local", "levies": [
            {"name": "NCD", "origin": "any", "asset": "any", "rate": 1, "per": {}},
            {"name": "VAT", "origin": "any", "asset": "any", "rate": 7.5, "per": {}, "vat": True}]},
        "ng_foreign": {"name": "Nigeria - foreign", "levies": [
            {"name": "NCD", "origin": "any", "asset": "any", "rate": 1, "per": {}},
            {"name": "Cabotage", "origin": "any", "asset": "vessel", "rate": 2, "per": {}},
            {"name": "Withholding Tax", "origin": "foreign", "asset": "any", "rate": 0,
             "per": {"Arbeid": 5, "Materieel": 7, "Materiaal": 6, "Onderaanneming": 8}},
            {"name": "VAT", "origin": "any", "asset": "any", "rate": 7.5, "per": {}, "vat": True}]},
        "my_local": {"name": "Malaysia - local (template)", "levies": [
            {"name": "SST", "origin": "any", "asset": "any", "rate": 8, "per": {}, "vat": True}]},
        "my_foreign": {"name": "Malaysia - foreign (template)", "levies": [
            {"name": "Withholding Tax", "origin": "foreign", "asset": "any", "rate": 10, "per": {}},
            {"name": "SST", "origin": "any", "asset": "any", "rate": 8, "per": {}, "vat": True}]},
    }


# --------------------------------------------------------------------------- #
# model
# --------------------------------------------------------------------------- #
class Node:
    __slots__ = ("id", "num", "level", "desc", "mult", "group", "children",
                 "note", "aantal", "unit", "hours", "cost", "net",
                 "costlines", "middel_id", "uren_raw", "uurloon_code", "alt", "loc",
                 "base_cost", "base_net", "base_hours")

    def __init__(self):
        self.children = []
        self.costlines = []


class Model:
    """Parsed IBIS budget. ``nodes`` is a dict id->Node; ``top`` is the ordered
    list of first-level chapter ids. Baselines (``base_*``) hold the as-loaded
    figures so edits can be expressed as deltas for write-back."""

    def __init__(self):
        self.header = {}
        self.valutas = {}
        self.nodes = {}
        self.top = []
        self.wages = []
        self.staart_lines = []
        self.parent = {}

    # --- geometry ---------------------------------------------------------- #
    def eff_mult(self, nid):
        f, x = 1.0, self.parent.get(nid)
        while x is not None:
            f *= (self.nodes[x].mult or 1.0)
            x = self.parent.get(x)
        return f

    def leaves(self):
        return [n for n in self.nodes.values() if not n.group and n.costlines]


# --------------------------------------------------------------------------- #
# parse
# --------------------------------------------------------------------------- #
def _open_bytes(data: bytes):
    """SQLite needs a file; write the bytes to a temp path and open it. Returns
    (connection, tmp_path) so the caller can clean up."""
    fd, path = tempfile.mkstemp(suffix=".xtb")
    os.write(fd, data)
    os.close(fd)
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    return con, path


def _cols(con, table):
    return {r[1] for r in con.execute(f"PRAGMA table_info({table})")}


def load(data: bytes) -> Model:
    con, _tmp = _open_bytes(data)
    try:
        c = con.cursor()
        m = Model()
        h = c.execute("SELECT Naam,Versie,Datum,Valuta,NettoTotaal,BrutoTotaal,"
                      "TotaalUren,BrutoStaart FROM Begrotingen").fetchone()
        m.header = {
            "name": h["Naam"], "versie": h["Versie"], "datum": h["Datum"],
            "calc_ccy": h["Valuta"], "hours": r2(h["TotaalUren"]),
            "stored_cost": r2(h["NettoTotaal"]), "stored_net": r2(h["BrutoTotaal"]),
            "stored_staart": r2(h["BrutoStaart"]),
        }
        for v in c.execute("SELECT Valuta,Teken,Koers FROM BegrotingValutas"):
            m.valutas[v["Valuta"]] = {"teken": v["Teken"], "koers": v["Koers"]}

        regels, children = {}, {}
        for r in c.execute("SELECT Id,ParentId,Regelnummer,Type,Multipliciteit,"
                            "Omschrijving FROM BegrotingsRegels"):
            regels[r["Id"]] = r
            children.setdefault(r["ParentId"], []).append(r["Id"])
        for arr in children.values():
            arr.sort(key=lambda i: regels[i]["Regelnummer"])

        kp = {r["Id"]: dict(r) for r in c.execute(
            "SELECT k.Id,k.MiddelId,k.Hoeveelheid,k.Uren,k.LangeTekst,"
            "k.NettoArbeid,k.NettoMateriaal,k.NettoMaterieel,k.NettoOnderaanneming,"
            "k.BrutoArbeid,k.BrutoMateriaal,k.BrutoMaterieel,k.BrutoOnderaanneming,"
            "m.Eenheid,m.UurloonCode,k.AlternatieveCode,k.LocatieCode "
            "FROM Kostenposten k LEFT JOIN Middelen m ON k.MiddelId=m.MiddelId")}
        el = {r["Id"]: dict(r) for r in c.execute(
            "SELECT Id,LangTekst,Hoeveelheid,Eenheid,TotaalUur,NettoTotaal,"
            "BrutoTotaal FROM Elementen")}

        m.wages = [{"id": w["Id"], "code": w["UurloonCode"], "desc": w["Omschrijving"],
                    "rate": w["Bedrag"], "hours": w["TotaalUren"]}
                   for w in c.execute("SELECT Id,UurloonCode,Omschrijving,Bedrag,"
                                       "TotaalUren FROM UurloonBedragen ORDER BY Id")]
        m.staart_lines = [{"id": s["Id"], "vol": s["Volgnummer"], "name": s["Omschrijving"],
                           "func": s["FunctieSoort"], "pct": s["Waarde"], "kost": s["Kostensoort"]}
                          for s in c.execute(
                              "SELECT Id,Volgnummer,Omschrijving,FunctieSoort,Waarde,"
                              "Kostensoort FROM BegrotingBladen "
                              "WHERE FunctieSoort IN ('TOP','VBH') ORDER BY Volgnummer")]

        def costlines(rid):
            k = kp.get(rid)
            if not k:
                return []
            out = []
            for nk, bk, typ, cat in CATS:
                net = k[nk] or 0
                bru = k[bk] or 0
                if abs(net) < 1e-9 and abs(bru) < 1e-9:
                    continue
                mk = (bru / net - 1) * 100 if abs(net) > 1e-9 else None
                out.append({"type": typ, "cat": cat,
                            "hours": r2(k["Uren"]) if (typ == "Personnel" and k["Uren"]) else None,
                            "cost": r2(net),
                            "markup": None if mk is None else round(mk * 10) / 10,
                            "net": r2(bru), "unit": None})
            return out

        def build(rid, level, number):
            r = regels[rid]
            is_g = (r["Type"] == 0)
            n = Node()
            n.id, n.num, n.level = rid, number, level
            n.desc = r["Omschrijving"] or ""
            mult = r["Multipliciteit"]
            n.mult = 1.0 if mult is None else round(mult * 10000) / 10000
            n.group = is_g
            n.middel_id = None
            n.uren_raw = 0
            n.uurloon_code = None
            n.alt = n.loc = ""
            if is_g:
                e = el.get(rid) or {}
                n.note = (e["LangTekst"] or "").strip() if e.get("LangTekst") else ""
                n.aantal = e["Hoeveelheid"] if e.get("Hoeveelheid") is not None else None
                n.unit = e["Eenheid"] if e else None
                n.hours = r2(e["TotaalUur"]) if e.get("TotaalUur") else None
                n.cost = r2(e["NettoTotaal"]) if e.get("NettoTotaal") is not None else None
                n.net = r2(e["BrutoTotaal"]) if e.get("BrutoTotaal") is not None else None
            else:
                k = kp.get(rid) or {}
                n.note = (k["LangeTekst"] or "").strip() if k.get("LangeTekst") else ""
                n.aantal = k["Hoeveelheid"] if k.get("Hoeveelheid") is not None else None
                n.unit = k["Eenheid"] if k else None
                n.middel_id = k["MiddelId"]
                n.uren_raw = k["Uren"] or 0
                n.uurloon_code = k["UurloonCode"]
                n.alt = (k["AlternatieveCode"] or "").strip()
                n.loc = (k["LocatieCode"] or "").strip()
                cls = costlines(rid)
                n.costlines = cls
                n.hours = cls[0]["hours"] if (cls and cls[0]["type"] == "Personnel") else None
                n.cost = r2(sum(c["cost"] for c in cls)) if cls else 0
                n.net = r2(sum(c["net"] for c in cls)) if cls else 0
            m.nodes[rid] = n
            ci = 0
            for cid in children.get(rid, []):
                ci += 1
                child_num = (number + "." + str(ci)) if number else str(ci)
                n.children.append(cid)
                build(cid, level + 1, child_num)

        root_id = next((r["Id"] for r in regels.values() if r["ParentId"] is None), None)
        i = 0
        for cid in children.get(root_id, []):
            i += 1
            build(cid, 0, str(i))
            m.top.append(cid)

        _build_parents(m)
        _snapshot(m)
        return m
    finally:
        con.close()
        if _tmp and os.path.exists(_tmp):
            os.remove(_tmp)


def _build_parents(m: Model):
    m.parent = {}
    for t in m.top:
        m.parent[t] = None
        stack = [t]
        while stack:
            pid = stack.pop()
            for cid in m.nodes[pid].children:
                m.parent[cid] = pid
                stack.append(cid)


def _snapshot(m: Model):
    """Freeze baselines and derive per-cost-line unit prices (cost / aantal)."""
    for n in m.nodes.values():
        n.base_cost, n.base_net, n.base_hours = n.cost, n.net, n.hours
        if not n.group:
            a = n.aantal
            has_a = a is not None and abs(a) > 1e-9
            for cl in n.costlines:
                cl["unit"] = (cl["cost"] / a) if has_a else cl["cost"]


# --------------------------------------------------------------------------- #
# edits + recompute
# --------------------------------------------------------------------------- #
def apply_edits(m: Model, edits: dict | None):
    """``edits`` maps leaf id -> {aantal, unit:{cat:price}, markup:{cat:pct}}.
    Mutates the model's leaf costlines and re-rolls groups + header from
    baselines. Idempotent: call with the full edit set each time."""
    edits = edits or {}
    # reset groups + leaves to baseline, then re-apply
    for n in m.nodes.values():
        if n.group:
            n.cost, n.net, n.hours = n.base_cost, n.base_net, n.base_hours
    for nid, n in m.nodes.items():
        if n.group:
            continue
        e = edits.get(str(nid)) or edits.get(nid) or {}
        if "aantal" in e and e["aantal"] is not None:
            n.aantal = e["aantal"]
        for cl in n.costlines:
            u = (e.get("unit") or {}).get(cl["cat"])
            if u is not None and cl["type"] != "Personnel":
                cl["unit"] = u
            mk = (e.get("markup") or {}).get(cl["cat"])
            if mk is not None:
                cl["markup"] = mk
        _recalc_leaf(m, n)
    _reroll(m, edits)


def _recalc_leaf(m: Model, n: Node):
    a = n.aantal
    has_a = a is not None and abs(a) > 1e-9
    hpa = (n.uren_raw / a) if has_a else 0
    for cl in n.costlines:
        cl["cost"] = r2(cl["unit"] * a) if has_a else r2(cl["unit"])
        f = 1.0 if cl["markup"] is None else (1 + cl["markup"] / 100)
        cl["net"] = r2(cl["cost"] * f)
        if cl["type"] == "Personnel":
            cl["hours"] = r2(hpa * a) if has_a else n.uren_raw
    n.cost = r2(sum(c["cost"] for c in n.costlines))
    n.net = r2(sum(c["net"] for c in n.costlines))
    p = next((c for c in n.costlines if c["type"] == "Personnel"), None)
    n.hours = p["hours"] if p else None


def _reroll(m: Model, edits):
    for nid, n in m.nodes.items():
        if n.group:
            continue
        d_cost = r2(n.cost - n.base_cost)
        d_net = r2(n.net - n.base_net)
        d_h = r2((n.hours or 0) - (n.base_hours or 0))
        if abs(d_cost) < 1e-9 and abs(d_net) < 1e-9 and abs(d_h) < 1e-9:
            continue
        f, x = 1.0, m.parent.get(nid)
        while x is not None:
            f *= (m.nodes[x].mult or 1.0)
            g = m.nodes[x]
            g.cost = r2((g.cost or 0) + d_cost * f)
            g.net = r2((g.net or 0) + d_net * f)
            g.hours = r2((g.hours or 0) + d_h * f)
            x = m.parent.get(x)


# --------------------------------------------------------------------------- #
# staart
# --------------------------------------------------------------------------- #
def staart_top(m: Model, override=None):
    """Return the effective list of TOP staart rows as {name,pct,kost}. If
    ``override`` (list of {name,pct}) is given, it replaces the file staart."""
    if override is not None:
        return [{"name": s.get("name", ""), "pct": s.get("pct", 0), "kost": s.get("kost", "AKS")}
                for s in override]
    return [{"name": s["name"], "pct": s["pct"], "kost": s.get("kost")}
            for s in m.staart_lines if s["func"] == "TOP"]


def staart_rate(top_rows):
    f = 1.0
    for s in top_rows:
        f *= (1 + (s["pct"] or 0))
    return f - 1


# --------------------------------------------------------------------------- #
# quote (levy + VAT) engine  -- faithful port of qCompute
# --------------------------------------------------------------------------- #
def _classify(l, tok_foreign="foreign", tok_vessel="vsl"):
    hay = ((l["desc"] or "") + " " + l["loc"] + " " + l["alt"]).lower()
    if l["loc"]:
        origin = "Foreign" if l["loc"].lower() == tok_foreign else "Local"
    else:
        origin = "Foreign" if "foreign" in hay else "Local"
    if l["alt"]:
        vessel = (l["alt"].lower() == tok_vessel)
    else:
        vessel = "vessel" in hay
    return {"origin": origin, "vessel": vessel}


def _q_leaves(m: Model):
    out = []
    for n in m.leaves():
        cost = {c: 0.0 for c in CATKEYS}
        sales = {c: 0.0 for c in CATKEYS}
        for cl in n.costlines:
            cost[cl["cat"]] += cl["cost"] or 0
            sales[cl["cat"]] += cl["net"] or 0
        out.append({"id": n.id, "desc": n.desc, "alt": n.alt, "loc": n.loc,
                    "aantal": n.aantal, "cost": cost, "sales": sales})
    return out


def _match(lv, cl):
    return ((lv["origin"] == "any" or lv["origin"].lower() == cl["origin"].lower())
            and (lv["asset"] == "any"
                 or (lv["asset"] == "vessel" and cl["vessel"])
                 or (lv["asset"] == "nonvessel" and not cl["vessel"])))


def _rate(lv, cat):
    p = (lv.get("per") or {}).get(cat)
    if p is not None and p != "":
        return float(p)
    return float(lv.get("rate") or 0)


def compute(m: Model, edits=None, staart_override=None, levies=None,
            tok_foreign="foreign", tok_vessel="vsl"):
    """The full roll-up + staart + levy/VAT quote. Returns the header cards and
    per-line detail. ``levies`` is an ordered list; VAT rows carry ``vat=True``
    and apply on the running total, others compound per matching cost type."""
    apply_edits(m, edits)
    top = staart_top(m, staart_override)
    sr = staart_rate(top)
    levies = levies or []

    leaves = _q_leaves(m)
    cells = []
    for l in leaves:
        mult = m.eff_mult(l["id"])
        cl = _classify(l, tok_foreign, tok_vessel)
        for cat in CATKEYS:
            sell = l["sales"].get(cat, 0)
            cst = l["cost"].get(cat, 0)
            if abs(sell) > 1e-9 or abs(cst) > 1e-9:
                cells.append({"l": l, "cl": cl, "cat": cat, "cost": cst, "sell": sell,
                              "staart": sell * sr, "val": sell * (1 + sr), "mult": mult})

    sells_base = sum(x["sell"] * x["mult"] for x in cells)
    raw_cost = sum(x["cost"] * x["mult"] for x in cells)
    staart_amt = sum(x["staart"] * x["mult"] for x in cells)
    gross = sells_base + staart_amt

    per_line = {l["id"]: {"levy": {}, "cost": 0, "sell": 0, "staart": 0,
                          "gross": 0, "sub": 0, "vat": 0} for l in leaves}
    for x in cells:
        p = per_line[x["l"]["id"]]
        p["cost"] += x["cost"]
        p["sell"] += x["sell"]
        p["staart"] += x["staart"]
        p["gross"] += x["val"]

    steps, running, amt_by_idx = [], gross, {}
    idxd = list(enumerate(levies))
    for idx, lv in [(i, l) for i, l in idxd if not l.get("vat")]:
        amt = 0.0
        for x in cells:
            if not _match(lv, x["cl"]):
                continue
            r = _rate(lv, x["cat"]) / 100
            a = x["val"] * r
            x["val"] += a
            amt += a * x["mult"]
            per_line[x["l"]["id"]]["levy"][idx] = per_line[x["l"]["id"]]["levy"].get(idx, 0) + a
        running += amt
        amt_by_idx[idx] = amt
        steps.append({"lv": lv, "amt": amt, "running": running})
    for idx, lv in [(i, l) for i, l in idxd if l.get("vat")]:
        rt = (float(lv.get("rate") or 0)) / 100
        amt = running * rt
        for x in cells:
            per_line[x["l"]["id"]]["vat"] += x["val"] * rt
        running += amt
        amt_by_idx[idx] = amt
        steps.append({"lv": lv, "amt": amt, "running": running})
    for x in cells:
        per_line[x["l"]["id"]]["sub"] += x["val"]

    non_vat_tot = sum(s["amt"] for s in steps if not s["lv"].get("vat"))
    vat_amt = sum(s["amt"] for s in steps if s["lv"].get("vat"))
    excl = gross + non_vat_tot
    incl = excl + vat_amt

    return {
        "leaves": leaves, "cells": cells, "sr": sr, "top": top,
        "sells_base": sells_base, "raw_cost": raw_cost, "markup": sells_base - raw_cost,
        "staart_amt": staart_amt, "gross": gross, "steps": steps, "amt_by_idx": amt_by_idx,
        "per_line": per_line, "excl": excl, "vat_amt": vat_amt, "non_vat_tot": non_vat_tot,
        "levy_tot": non_vat_tot + vat_amt, "incl": incl,
    }


def header_cards(R):
    """The five top-of-page figures, in calc currency."""
    return {"cost": R["raw_cost"], "markup": R["markup"], "staart": R["staart_amt"],
            "levy": R["levy_tot"], "net": R["incl"]}


# --------------------------------------------------------------------------- #
# seed levies from the file's own staart
# --------------------------------------------------------------------------- #
def levies_from_staart(m: Model):
    lines = [x for x in m.staart_lines if x["func"] == "TOP" and (x["name"] or "").strip()]
    if not lines:
        return None
    by_vol = {}
    for x in lines:
        by_vol.setdefault(x["vol"], []).append(x)
    out = []
    for vol in sorted(by_vol):
        g = by_vol[vol]
        name = (g[0]["name"] or f"Levy {vol}").strip()
        lo = name.lower()
        lv = {"name": name, "origin": "any", "asset": "any", "rate": 0, "per": {}}
        if "vat" in lo or "btw" in lo:
            lv["vat"] = True
        if "cabotage" in lo:
            lv["asset"] = "vessel"
        if "wht" in lo or ("with" in lo and "tax" in lo):
            lv["origin"] = "foreign"
        if len(g) == 1 and (not g[0]["kost"] or g[0]["kost"] == "AKS"):
            lv["rate"] = r2((g[0]["pct"] or 0) * 100)
        else:
            for x in g:
                cat = QKMAP.get(x["kost"])
                if cat:
                    lv["per"][cat] = r2((x["pct"] or 0) * 100)
                else:
                    lv["rate"] = r2((x["pct"] or 0) * 100)
        out.append(lv)
    return out


# --------------------------------------------------------------------------- #
# report drill-down data
# --------------------------------------------------------------------------- #
def report_items(m: Model):
    items = []

    def walk(nid, parents):
        n = m.nodes[nid]
        if n.group:
            for c in n.children:
                walk(c, parents + [n.desc])
            return
        for cl in n.costlines:
            items.append({"cat": cl["cat"], "res": n.desc or "(line)", "net": cl["net"] or 0,
                          "cost": cl["cost"] or 0, "qty": n.aantal, "markup": cl["markup"],
                          "path": parents})
    for t in m.top:
        walk(t, [])
    return items


# --------------------------------------------------------------------------- #
# Excel export  (live formulas; Excel recalculates on open)
# --------------------------------------------------------------------------- #
def to_xlsx(m: Model, R, wage_rates=None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter as CL

    wage_rates = wage_rates or {w["code"]: w["rate"] for w in m.wages}
    st_r = R["sr"]
    non_vat = [(i, l) for i, l in enumerate(_ordered_levies(R)) if not l.get("vat")]
    vat_lv = next((l for _, l in enumerate(_ordered_levies(R)) if l.get("vat")), None)
    vat_dec = ((float(vat_lv.get("rate") or 0)) / 100) if vat_lv else 0

    # column layout (0-based)
    iS, iItem, iDesc, iCT, iOrg, iVes, iAant, iUnit, iCost, iMk, iSales, iStaart, iGross = range(13)
    iLevy0 = 13
    P = len(non_vat)
    iSub, iVat, iNet = iLevy0 + P, iLevy0 + P + 1, iLevy0 + P + 2
    NCOL = iNet + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote lines"
    head = ["S", "Item", "Description", "Cost type", "Origin", "Vessel", "Aantal",
            "Unit cost", "Cost", "Markup %", "Sales", "Staart", "Gross"]
    for _, lv in non_vat:
        head.append(lv["name"])
    head += ["Subtotal excl VAT", "VAT", "Net price"]

    bold = Font(name="Arial", bold=True)
    reg = Font(name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1f2937")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    for j, htxt in enumerate(head, start=1):
        cell = ws.cell(row=1, column=j, value=htxt)
        cell.font = hdr_font
        cell.fill = hdr_fill

    per_line = R["per_line"]
    q_leaf = {l["id"]: l for l in R["leaves"]}
    money_fmt = '#,##0.00'
    rows_written = [1]  # header row index
    item_no = [0]

    def eff_rate(lv, lf):
        cl = _classify(lf)
        if not _match(lv, cl):
            return 0
        return _rate(lv, _primary(lf)) / 100

    def pad(lvl):
        return "  " * max(0, lvl)

    def walk(nid):
        n = m.nodes[nid]
        if n.group:
            kids = [walk(c) for c in n.children]
            ri = ws.max_row + 1
            row = ws.cell(row=ri, column=iS + 1, value=n.level + 1)
            ws.cell(row=ri, column=iDesc + 1, value=pad(n.level) + n.desc).font = bold
            mval = n.mult if n.mult is not None else 1
            sc = abs(mval - 1) > 1e-9
            v = {}
            sumcols = [iCost, iSales, iStaart, iGross, iSub, iVat, iNet] + [iLevy0 + k for k in range(P)]
            for ci in sumcols:
                parts = [CL(ci + 1) + str(k["row"]) for k in kids if ci in k["v"]]
                cv = sum(k["v"].get(ci, 0) for k in kids) * mval
                v[ci] = cv
                if parts:
                    inner = "+".join(parts)
                    f = f"={mval}*({inner})" if sc else "=" + inner
                    cc = ws.cell(row=ri, column=ci + 1, value=f)
                    cc.number_format = money_fmt
            return {"row": ri, "v": v}
        lf = q_leaf.get(nid)
        if not lf:
            ri = ws.max_row + 1
            ws.cell(row=ri, column=iDesc + 1, value=pad(n.level) + n.desc)
            return {"row": ri, "v": {}}
        item_no[0] += 1
        pl = per_line[nid]
        cl = _classify(lf)
        cost, sales = pl["cost"], pl["sell"]
        mk = (sales / cost - 1) * 100 if cost > 1e-9 else 0
        unit = (cost / lf["aantal"]) if (lf["aantal"] and abs(lf["aantal"]) > 1e-9) else cost
        ri = ws.max_row + 1
        R_ = ri
        ws.cell(row=ri, column=iItem + 1, value=item_no[0])
        ws.cell(row=ri, column=iDesc + 1, value=pad(n.level) + n.desc)
        ws.cell(row=ri, column=iCT + 1, value=QLBL[_primary(lf)])
        ws.cell(row=ri, column=iOrg + 1, value=cl["origin"])
        ws.cell(row=ri, column=iVes + 1, value="Yes" if cl["vessel"] else "")
        ws.cell(row=ri, column=iAant + 1, value=lf["aantal"])
        ws.cell(row=ri, column=iUnit + 1, value=round(unit * 1e5) / 1e5)
        c_cost = ws.cell(row=ri, column=iCost + 1, value=f"={CL(iUnit+1)}{R_}*{CL(iAant+1)}{R_}")
        ws.cell(row=ri, column=iMk + 1, value=round(mk * 1e4) / 1e4)
        c_sales = ws.cell(row=ri, column=iSales + 1,
                          value=f"={CL(iCost+1)}{R_}*(1+{CL(iMk+1)}{R_}/100)")
        c_st = ws.cell(row=ri, column=iStaart + 1, value=f"={CL(iSales+1)}{R_}*{st_r}")
        c_gr = ws.cell(row=ri, column=iGross + 1, value=f"={CL(iSales+1)}{R_}+{CL(iStaart+1)}{R_}")
        v = {iCost: cost, iSales: sales, iStaart: pl["staart"], iGross: pl["gross"]}
        for k, (idx, lv) in enumerate(non_vat):
            col = iLevy0 + k
            er = eff_rate(lv, lf)
            ws.cell(row=ri, column=col + 1,
                    value=f"=SUM({CL(iGross+1)}{R_}:{CL(col)}{R_})*{er}")
            v[col] = pl["levy"].get(idx, 0)
        ws.cell(row=ri, column=iSub + 1,
                value=f"=SUM({CL(iGross+1)}{R_}:{CL(iLevy0+P)}{R_})")
        v[iSub] = pl["sub"]
        ws.cell(row=ri, column=iVat + 1, value=f"={CL(iSub+1)}{R_}*{vat_dec}")
        v[iVat] = pl["vat"]
        ws.cell(row=ri, column=iNet + 1, value=f"={CL(iSub+1)}{R_}+{CL(iVat+1)}{R_}")
        v[iNet] = pl["sub"] + pl["vat"]
        for ci in (iCost, iSales, iStaart, iGross, iSub, iVat, iNet, *[iLevy0 + k for k in range(P)]):
            ws.cell(row=ri, column=ci + 1).number_format = money_fmt
        return {"row": R_, "v": v}

    roots = [walk(t) for t in m.top]
    tr = ws.max_row + 1
    ws.cell(row=tr, column=iDesc + 1, value="TOTAL").font = bold
    for ci in [iCost, iSales, iStaart, iGross, iSub, iVat, iNet] + [iLevy0 + k for k in range(P)]:
        parts = [CL(ci + 1) + str(k["row"]) for k in roots if ci in k["v"]]
        if parts:
            c = ws.cell(row=tr, column=ci + 1, value="=" + "+".join(parts))
            c.number_format = money_fmt
            c.font = bold

    # hourly rates sheet
    wr = wb.create_sheet("Hourly rates")
    wr.cell(row=1, column=1, value=f"Hourly rates  {m.header['name']}").font = bold
    for j, htxt in enumerate(["Code", "Description", "Hours", "Hourly rate",
                              "Day rate (12h)", "Labour cost"], start=1):
        wr.cell(row=3, column=j, value=htxt).font = bold
    ri = 4
    for w in sorted(m.wages, key=lambda a: (a["desc"] or "")):
        rate = wage_rates.get(w["code"], w["rate"]) or 0
        hrs = w["hours"] or 0
        wr.cell(row=ri, column=1, value=w["code"])
        wr.cell(row=ri, column=2, value=w["desc"] or "")
        wr.cell(row=ri, column=3, value=hrs)
        wr.cell(row=ri, column=4, value=r2(rate)).number_format = money_fmt
        wr.cell(row=ri, column=5, value=f"=D{ri}*12").number_format = money_fmt
        wr.cell(row=ri, column=6, value=f"=C{ri}*D{ri}").number_format = money_fmt
        ri += 1

    # summary sheet
    sm = wb.create_sheet("Summary")
    sm.cell(row=1, column=1, value=f"Quote summary  {m.header['name']}").font = bold
    rows = [("Cost (raw)", R["raw_cost"]), ("Markup", R["markup"]),
            ("Net of line items (sales)", R["sells_base"]),
            ("Staart (overhead / profit / CAR)", R["staart_amt"]),
            ("Price (gross)", R["gross"]), ("", None)]
    for s in R["steps"]:
        if not s["lv"].get("vat"):
            rows.append((s["lv"]["name"], s["amt"]))
    rows += [("Totaal excl BTW", R["excl"]), ("VAT", R["vat_amt"]),
             ("Totaal incl BTW (quote)", R["incl"])]
    ri = 3
    for label, val in rows:
        sm.cell(row=ri, column=1, value=label)
        if val is not None:
            sm.cell(row=ri, column=2, value=r2(val)).number_format = money_fmt
        ri += 1

    # widths
    for col, w in {"A": 6, "B": 8, "C": 40}.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ordered_levies(R):
    return [s["lv"] for s in R["steps"]]


def _primary(lf):
    b, v = "Materieel", -1
    for c in CATKEYS:
        if (lf["sales"].get(c, 0)) > v:
            v = lf["sales"].get(c, 0)
            b = c
    return b


# --------------------------------------------------------------------------- #
# write amendments back into a NEW .xtb copy
# --------------------------------------------------------------------------- #
def write_xtb(orig: bytes, m: Model, edits=None, wage_rates=None,
              staart_override=None) -> bytes:
    """Return a new ``.xtb`` (bytes) with the edits applied: Kostenposten,
    Middelen, Elementen and Begroting roll-ups, wage rates, and - only if a
    staart override is supplied - a rebuilt BegrotingBladen block. The original
    is never modified. Bumps Versie and stamps Datum, matching IBIS behaviour."""
    apply_edits(m, edits)
    top = staart_top(m, staart_override)
    sr = staart_rate(top)
    wage_rates = wage_rates or {}

    fd, path = tempfile.mkstemp(suffix=".xtb")
    os.write(fd, orig)
    os.close(fd)
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    try:
        c = con.cursor()
        now = datetime.now(timezone.utc)
        today = now.strftime("%Y-%m-%d")
        stamp = now.strftime("%Y-%m-%dT%H:%M:%S")

        elem_delta = {}
        beg = {"net": {k: 0.0 for k in CATKEYS}, "h": 0.0}
        mid_cols = _cols(con, "Middelen")
        el_cols = _cols(con, "Elementen")

        for nid, n in m.nodes.items():
            if n.group:
                continue
            e = (edits or {}).get(str(nid)) or (edits or {}).get(nid)
            if not e:
                continue  # unchanged leaf
            cnet = {k: 0.0 for k in CATKEYS}
            cbru = {k: 0.0 for k in CATKEYS}
            for cl in n.costlines:
                cnet[cl["cat"]] = cl["cost"]
                cbru[cl["cat"]] = cl["net"]
            onet = {k: 0.0 for k in CATKEYS}
            obru = {k: 0.0 for k in CATKEYS}
            # baseline costlines rebuilt from base_cost is not stored per-cat; recompute
            # from the model's *original* by reloading is avoided - instead we track
            # deltas via node baselines at the cost-type sum level using stored cost.
            # For per-cat write-back we use current minus (unit@baseline * base aantal).
            for cl in n.costlines:
                # baseline cat cost = original stored (before edit) -> reconstruct
                pass
            # per-cat original: cost line 'unit' already reflects edits; use baseline sums
            # Simpler + exact: derive original per-cat from the DB row.
            row = con.execute(
                "SELECT NettoArbeid,NettoMaterieel,NettoMateriaal,NettoOnderaanneming,"
                "BrutoArbeid,BrutoMaterieel,BrutoMateriaal,BrutoOnderaanneming,Uren "
                "FROM Kostenposten WHERE Id=?", (nid,)).fetchone()
            onet = {"Arbeid": row["NettoArbeid"] or 0, "Materieel": row["NettoMaterieel"] or 0,
                    "Materiaal": row["NettoMateriaal"] or 0,
                    "Onderaanneming": row["NettoOnderaanneming"] or 0}
            obru = {"Arbeid": row["BrutoArbeid"] or 0, "Materieel": row["BrutoMaterieel"] or 0,
                    "Materiaal": row["BrutoMateriaal"] or 0,
                    "Onderaanneming": row["BrutoOnderaanneming"] or 0}
            dnet = {k: r5(cnet[k] - onet[k]) for k in CATKEYS}
            dbru = {k: r5(cbru[k] - obru[k]) for k in CATKEYS}
            cur_h = n.hours or 0
            d_h = r5(cur_h - (row["Uren"] or 0))
            netto_tot = r5(sum(cnet.values()))
            bruto_tot = r5(sum(cbru.values()))
            a = n.aantal or 0
            eprijs = r5(bruto_tot / a) if (a and abs(a) > 1e-9) else bruto_tot
            c.execute(
                "UPDATE Kostenposten SET Hoeveelheid=?,Uren=?,NettoArbeid=?,NettoMateriaal=?,"
                "NettoMaterieel=?,NettoOnderaanneming=?,BrutoArbeid=?,BrutoMateriaal=?,"
                "BrutoMaterieel=?,BrutoOnderaanneming=?,NettoTotaal=?,BrutoTotaal=?,"
                "Eenheidsprijs=? WHERE Id=?",
                (a, cur_h, r5(cnet["Arbeid"]), r5(cnet["Materiaal"]), r5(cnet["Materieel"]),
                 r5(cnet["Onderaanneming"]), r5(cbru["Arbeid"]), r5(cbru["Materiaal"]),
                 r5(cbru["Materieel"]), r5(cbru["Onderaanneming"]), netto_tot, bruto_tot,
                 eprijs, nid))

            if n.middel_id is not None:
                parts, vals = [], []
                ef = {"Materiaal": "EenheidsprijsMateriaal", "Materieel": "EenheidsprijsMaterieel",
                      "Onderaanneming": "EenheidsprijsOnderaanneming"}
                for cl in n.costlines:
                    col = ef.get(cl["cat"])
                    if col and col in mid_cols:
                        parts.append(f"{col}=?")
                        vals.append(r5(cl["cost"] / a) if (a and abs(a) > 1e-9) else r5(cl["cost"]))
                    fcol = "Factor" + cl["cat"]
                    if fcol in mid_cols:
                        parts.append(f"{fcol}=?")
                        vals.append(r5(cl["net"] / cl["cost"]) if cl["cost"] > 1e-9 else 1)
                pers = next((cl for cl in n.costlines if cl["cat"] == "Arbeid"), None)
                if pers and "NormUren" in mid_cols:
                    parts.append("NormUren=?")
                    vals.append(r5(cur_h / a) if (a and abs(a) > 1e-9) else cur_h)
                if "WijzigDatum" in mid_cols:
                    parts.append("WijzigDatum=?")
                    vals.append(today)
                if parts:
                    vals.append(n.middel_id)
                    c.execute("UPDATE Middelen SET " + ",".join(parts) + " WHERE MiddelId=?", vals)

            # propagate deltas up to element parents + begroting
            f, x = 1.0, m.parent.get(nid)
            while x is not None:
                f *= (m.nodes[x].mult or 1.0)
                d = elem_delta.setdefault(x, {"net": {k: 0.0 for k in CATKEYS},
                                              "bru": {k: 0.0 for k in CATKEYS}, "h": 0.0})
                for k in CATKEYS:
                    d["net"][k] += dnet[k] * f
                    d["bru"][k] += dbru[k] * f
                d["h"] += d_h * f
                x = m.parent.get(x)
            em = m.eff_mult(nid)
            for k in CATKEYS:
                beg["net"][k] += dbru[k] * em
            beg["h"] += d_h * em

        for eid, d in elem_delta.items():
            d_net_tot = sum(d["net"].values())
            d_bru_tot = sum(d["bru"].values())
            extra = ""
            extra_vals = []
            if "BtwGrondslagHoogArbeid" in el_cols:
                extra = (",BtwGrondslagHoogArbeid=BtwGrondslagHoogArbeid+?,"
                         "BtwGrondslagHoogMateriaal=BtwGrondslagHoogMateriaal+?,"
                         "BtwGrondslagHoogMaterieel=BtwGrondslagHoogMaterieel+?,"
                         "BtwGrondslagHoogOnderaanneming=BtwGrondslagHoogOnderaanneming+?,"
                         "BtwGrondslagHoogTotaal=BtwGrondslagHoogTotaal+?")
                extra_vals = [r5(d["bru"]["Arbeid"]), r5(d["bru"]["Materiaal"]),
                              r5(d["bru"]["Materieel"]), r5(d["bru"]["Onderaanneming"]),
                              r5(d_bru_tot)]
            c.execute(
                "UPDATE Elementen SET NettoArbeid=NettoArbeid+?,NettoMateriaal=NettoMateriaal+?,"
                "NettoMaterieel=NettoMaterieel+?,NettoOnderaanneming=NettoOnderaanneming+?,"
                "BrutoArbeid=BrutoArbeid+?,BrutoMateriaal=BrutoMateriaal+?,"
                "BrutoMaterieel=BrutoMaterieel+?,BrutoOnderaanneming=BrutoOnderaanneming+?" + extra +
                ",NettoTotaal=NettoTotaal+?,BrutoTotaal=BrutoTotaal+?,TotaalUur=TotaalUur+? "
                "WHERE Id=?",
                [r5(d["net"]["Arbeid"]), r5(d["net"]["Materiaal"]), r5(d["net"]["Materieel"]),
                 r5(d["net"]["Onderaanneming"]), r5(d["bru"]["Arbeid"]), r5(d["bru"]["Materiaal"]),
                 r5(d["bru"]["Materieel"]), r5(d["bru"]["Onderaanneming"])] + extra_vals +
                [r5(d_net_tot), r5(d_bru_tot), r5(d["h"]), eid])
            c.execute("UPDATE Elementen SET Eenheidsprijs = CASE WHEN Hoeveelheid IS NULL "
                      "OR Hoeveelheid=0 OR Hoeveelheid=1 THEN BrutoTotaal "
                      "ELSE BrutoTotaal/Hoeveelheid END WHERE Id=?", (eid,))

        b_sell = sum(beg["net"].values())
        c.execute(
            "UPDATE Begrotingen SET NettoArbeid=NettoArbeid+?,NettoMateriaal=NettoMateriaal+?,"
            "NettoMaterieel=NettoMaterieel+?,NettoOnderaanneming=NettoOnderaanneming+?,"
            "BrutoArbeid=BrutoArbeid+?,BrutoMateriaal=BrutoMateriaal+?,"
            "BrutoMaterieel=BrutoMaterieel+?,BrutoOnderaanneming=BrutoOnderaanneming+?,"
            "NettoTotaal=NettoTotaal+?,TotaalUren=TotaalUren+?,Versie=Versie+1,Datum=?",
            (r5(beg["net"]["Arbeid"]), r5(beg["net"]["Materiaal"]), r5(beg["net"]["Materieel"]),
             r5(beg["net"]["Onderaanneming"]), r5(beg["net"]["Arbeid"]), r5(beg["net"]["Materiaal"]),
             r5(beg["net"]["Materieel"]), r5(beg["net"]["Onderaanneming"]), r5(b_sell),
             r5(beg["h"]), stamp))

        for w in m.wages:
            nr = wage_rates.get(w["code"])
            if nr is not None and abs(nr - (w["rate"] or 0)) > 1e-9:
                c.execute("UPDATE UurloonBedragen SET Bedrag=? WHERE Id=?", (r5(nr), w["id"]))

        c.execute("UPDATE Begrotingen SET BrutoStaart=NettoTotaal*?, BrutoTotaal=NettoTotaal*(1+?)",
                  (sr, sr))

        if staart_override is not None:
            _write_staart(con, top)

        con.commit()
        with open(path, "rb") as fh:
            out = fh.read()
        return out
    finally:
        con.close()
        if os.path.exists(path):
            os.remove(path)


def _write_staart(con, top_rows):
    c = con.cursor()
    bid = con.execute("SELECT Id FROM Begrotingen LIMIT 1").fetchone()["Id"]
    sec, afd = -26, 1
    ex = con.execute("SELECT SectieId,Afdrukken FROM BegrotingBladen LIMIT 1").fetchone()
    if ex:
        sec, afd = ex["SectieId"], ex["Afdrukken"]
    c.execute("DELETE FROM BegrotingBladen WHERE BegrotingId=?", (bid,))
    row = con.execute("SELECT COALESCE(MAX(Id),0) m FROM BegrotingBladen").fetchone()
    idc = [row["m"]]
    vol = [0]

    def ins(func, oms, kost, waarde):
        idc[0] += 1
        c.execute("INSERT INTO BegrotingBladen (Id,BegrotingId,SectieId,Volgnummer,"
                  "CalculatieCode,Omschrijving,FunctieSoort,MeetstaatKoppelingId,Afdrukken,"
                  "Waarde,Kostensoort,Verzamelpost,Eenheid) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                  (idc[0], bid, sec, vol[0], "", oms, func, 0, afd, waarde, kost, "", ""))

    ins("TOT", "", "", 0)
    vol[0] += 1
    groups, seen = [], {}
    for s in top_rows:
        k = s.get("name", "") or ""
        if k not in seen:
            seen[k] = len(groups)
            groups.append({"name": k, "rows": []})
        groups[seen[k]]["rows"].append(s)
    for g in groups:
        ins("CMT", "", "", 0)
        vol[0] += 1
        for r in g["rows"]:
            ins("TOP", g["name"], r.get("kost") or "AKS", float(r.get("pct") or 0))
        vol[0] += 1
        ins("TTL", "", "", 0)
        vol[0] += 1
